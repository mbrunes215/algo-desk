#!/usr/bin/env python3
"""
Excel Report Generator for Algo Trading Desk

Reads from SQLite outcome tracker database and generates formatted Excel dashboard.
"""

import sqlite3
import argparse
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_db_path(db_arg):
    """Resolve database path, defaulting to trading.db in parent directory."""
    if db_arg:
        return db_arg
    algo_desk_dir = Path(__file__).parent.parent
    return str(algo_desk_dir / "trading.db")


def get_active_signals(db_path):
    """Get all PENDING signals sorted by edge descending."""
    columns = ["Date", "Ticker", "Direction", "Model%", "Market%",
               "Edge(bps)", "Confidence", "Size", "Metric", "Model Source"]
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT logged_at, ticker, direction, model_prob, market_prob,
                       edge_bps, confidence, position_size, weather_metric, model_source
                FROM signals
                WHERE status = 'PENDING'
                ORDER BY edge_bps DESC
            """)
            return columns, cursor.fetchall()
    except sqlite3.OperationalError:
        return columns, []


def get_trade_history(db_path):
    """Get all settled trades sorted by settle date descending."""
    columns = ["Signal Date", "Settle Date", "Ticker", "Direction",
               "Entry%", "Exit%", "Edge(bps)", "Size", "Outcome",
               "P&L(¢/contract)", "Total P&L(¢)", "City"]
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT s.logged_at, o.settled_at, s.ticker, s.direction,
                       s.market_prob, o.exit_price_cents, s.edge_bps,
                       s.position_size, o.outcome, o.pnl_cents,
                       s.weather_metric
                FROM signals s
                INNER JOIN outcomes o ON s.id = o.signal_id
                ORDER BY o.settled_at DESC
            """)
            return columns, cursor.fetchall()
    except sqlite3.OperationalError:
        return columns, []


def get_performance_summary(db_path):
    """Get summary statistics for dashboard."""
    default = {
        "total_signals": 0,
        "pending": 0,
        "wins": 0,
        "losses": 0,
        "win_rate": 0.0,
        "total_pnl": 0.0,
        "cities": {},
        "directions": {}
    }

    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()

            cursor.execute("SELECT COUNT(*) FROM signals")
            total_signals = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM signals WHERE status = 'PENDING'")
            pending = cursor.fetchone()[0]

            cursor.execute("""
                SELECT outcome, COUNT(*), SUM(pnl_cents)
                FROM outcomes
                GROUP BY outcome
            """)
            outcome_stats = cursor.fetchall()

            wins = losses = total_pnl = 0.0
            for outcome, count, pnl in outcome_stats:
                if outcome == "WIN":
                    wins = count
                    total_pnl += pnl if pnl else 0.0
                elif outcome == "LOSS":
                    losses = count
                    total_pnl += pnl if pnl else 0.0

            win_rate = wins / (wins + losses) if (wins + losses) > 0 else 0.0

            # By city
            cursor.execute("""
                SELECT s.weather_metric, COUNT(*), SUM(o.pnl_cents),
                       COUNT(CASE WHEN o.outcome = 'WIN' THEN 1 END)
                FROM signals s
                LEFT JOIN outcomes o ON s.id = o.signal_id
                WHERE s.status IN ('WIN', 'LOSS')
                GROUP BY s.weather_metric
                ORDER BY COUNT(*) DESC
            """)
            city_rows = cursor.fetchall()

            cities = {}
            for metric, count, pnl, city_wins in city_rows:
                city = metric.split("_")[0] if metric else "UNKNOWN"
                city_pnl = pnl if pnl else 0.0
                city_win_rate = city_wins / count if count > 0 else 0.0
                cities[city] = {
                    "signals": count,
                    "wins": city_wins,
                    "losses": count - city_wins,
                    "win_rate": city_win_rate,
                    "pnl": city_pnl
                }

            # By direction
            directions = {}
            for direction in ["BUY", "SELL"]:
                cursor.execute("""
                    SELECT COUNT(*), SUM(o.pnl_cents),
                           COUNT(CASE WHEN o.outcome = 'WIN' THEN 1 END)
                    FROM signals s
                    LEFT JOIN outcomes o ON s.id = o.signal_id
                    WHERE s.direction = ? AND s.status IN ('WIN', 'LOSS')
                """, (direction,))
                row = cursor.fetchone()
                count, pnl, dir_wins = (row[0] or 0, row[1] or 0.0, row[2] or 0)
                dir_win_rate = dir_wins / count if count > 0 else 0.0
                directions[direction] = {
                    "signals": count,
                    "wins": dir_wins,
                    "losses": count - dir_wins,
                    "win_rate": dir_win_rate,
                    "pnl": pnl
                }

            return {
                "total_signals": total_signals,
                "pending": pending,
                "wins": wins,
                "losses": losses,
                "win_rate": win_rate,
                "total_pnl": total_pnl,
                "cities": cities,
                "directions": directions
            }
    except sqlite3.OperationalError:
        return default


def format_header(ws, row, columns):
    """Format header row with dark gray background and white text."""
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, text in enumerate(columns, 1):
        cell = ws.cell(row, col, text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def format_generated_timestamp(ws, timestamp):
    """Add generated timestamp at top of sheet."""
    ws.insert_rows(1)
    ws['A1'] = f"Generated: {timestamp}"
    ws['A1'].font = Font(name="Arial", size=9, italic=True)


def set_auto_width(ws, columns):
    """Set reasonable column widths based on content."""
    for col, col_name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_active_signals_sheet(wb, db_path, timestamp):
    """Create Active Signals sheet."""
    ws = wb.create_sheet("Active Signals", 0)

    columns, data = get_active_signals(db_path)
    format_generated_timestamp(ws, timestamp)

    format_header(ws, 2, columns)

    buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    sell_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    body_font = Font(name="Arial", size=10)

    for row_idx, row_data in enumerate(data, 3):
        direction = row_data[2]
        row_fill = buy_fill if direction == "BUY" else sell_fill

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = body_font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

            if col_idx in [1, 4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"
    set_auto_width(ws, columns)
    return ws


def create_trade_history_sheet(wb, db_path, timestamp):
    """Create Trade History sheet with running totals."""
    ws = wb.create_sheet("Trade History", 1)

    columns, data = get_trade_history(db_path)
    format_generated_timestamp(ws, timestamp)

    format_header(ws, 2, columns)

    win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loss_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)

    win_count = loss_count = total_pnl = 0

    for row_idx, row_data in enumerate(data, 3):
        outcome = row_data[8]
        pnl = row_data[9]
        row_fill = win_fill if outcome == "WIN" else loss_fill

        if outcome == "WIN":
            win_count += 1
        else:
            loss_count += 1
        total_pnl += pnl

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = body_font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    total_row = len(data) + 3
    ws.cell(total_row, 1, "TOTALS")
    ws.cell(total_row, 9, f"Wins: {win_count} | Losses: {loss_count} | Win Rate: {win_count/(win_count+loss_count)*100:.1f}%" if (win_count+loss_count) > 0 else "N/A")
    ws.cell(total_row, 11, f"Total P&L: {total_pnl:.1f}¢")

    for col in range(1, 12):
        ws.cell(total_row, col).font = bold_font
        ws.cell(total_row, col).fill = total_fill

    set_auto_width(ws, columns)
    return ws


def create_performance_summary_sheet(wb, db_path, timestamp):
    """Create Performance Summary dashboard sheet."""
    ws = wb.create_sheet("Performance Summary", 2)

    summary = get_performance_summary(db_path)
    format_generated_timestamp(ws, timestamp)

    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)

    row = 3

    ws.cell(row, 1, "OVERALL STATS").font = Font(name="Arial", size=12, bold=True)
    row += 1

    stats = [
        ("Total Signals", summary["total_signals"]),
        ("Pending Signals", summary["pending"]),
        ("Total Wins", int(summary["wins"])),
        ("Total Losses", int(summary["losses"])),
        ("Win Rate", f"{summary['win_rate']*100:.1f}%"),
        ("Total P&L", f"{summary['total_pnl']:.1f}¢")
    ]

    for label, value in stats:
        ws.cell(row, 1, label).font = bold_font
        ws.cell(row, 2, value).font = body_font
        row += 1

    row += 1
    ws.cell(row, 1, "BY CITY").font = Font(name="Arial", size=12, bold=True)
    row += 1

    city_headers = ["City", "Signals", "Wins", "Losses", "Win Rate", "P&L"]
    for col, header in enumerate(city_headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1

    for city, stats_dict in sorted(summary["cities"].items()):
        ws.cell(row, 1, city).font = body_font
        ws.cell(row, 2, stats_dict["signals"]).font = body_font
        ws.cell(row, 3, stats_dict["wins"]).font = body_font
        ws.cell(row, 4, stats_dict["losses"]).font = body_font
        ws.cell(row, 5, f"{stats_dict['win_rate']*100:.1f}%").font = body_font
        ws.cell(row, 6, f"{stats_dict['pnl']:.1f}¢").font = body_font
        row += 1

    row += 1
    ws.cell(row, 1, "BY DIRECTION").font = Font(name="Arial", size=12, bold=True)
    row += 1

    dir_headers = ["Direction", "Signals", "Wins", "Losses", "Win Rate", "P&L"]
    for col, header in enumerate(dir_headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1

    for direction in ["BUY", "SELL"]:
        if direction in summary["directions"]:
            stats_dict = summary["directions"][direction]
            ws.cell(row, 1, direction).font = body_font
            ws.cell(row, 2, stats_dict["signals"]).font = body_font
            ws.cell(row, 3, stats_dict["wins"]).font = body_font
            ws.cell(row, 4, stats_dict["losses"]).font = body_font
            ws.cell(row, 5, f"{stats_dict['win_rate']*100:.1f}%").font = body_font
            ws.cell(row, 6, f"{stats_dict['pnl']:.1f}¢").font = body_font
            row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    return ws


def main():
    parser = argparse.ArgumentParser(description="Generate Excel report from trading database")
    parser.add_argument("--db", help="Path to trading.db (default: trading.db in parent dir)")
    parser.add_argument("--output", help="Output Excel file path")
    args = parser.parse_args()

    db_path = get_db_path(args.db)

    if not Path(db_path).exists():
        print(f"Error: Database not found at {db_path}")
        return 1

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()
    wb.remove(wb.active)

    create_active_signals_sheet(wb, db_path, timestamp)
    create_trade_history_sheet(wb, db_path, timestamp)
    create_performance_summary_sheet(wb, db_path, timestamp)

    output_path = args.output or str(Path(__file__).resolve().parent.parent / "Trading Dashboard.xlsx")

    wb.save(output_path)

    print(f"Report generated successfully!")
    print(f"Output: {output_path}")
    print(f"Timestamp: {timestamp}")

    return 0


def generate_report(db_path: str = None, output_path: str = None) -> str:
    """
    Programmatic entry point for use by main.py.

    Args:
        db_path: Path to trading.db (default: auto-resolve)
        output_path: Where to save the .xlsx (default: Trading Dashboard.xlsx next to algo-desk)

    Returns:
        Path where the file was saved.
    """
    db_path = db_path or get_db_path(None)
    output_path = output_path or str(Path(__file__).resolve().parent.parent / "Trading Dashboard.xlsx")

    if not Path(db_path).exists():
        raise FileNotFoundError(f"Database not found: {db_path}")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()
    wb.remove(wb.active)

    create_active_signals_sheet(wb, db_path, timestamp)
    create_trade_history_sheet(wb, db_path, timestamp)
    create_performance_summary_sheet(wb, db_path, timestamp)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    exit(main())
